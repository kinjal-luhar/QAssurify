"""
Reporter utility for QA Agent
Handles logging test results to console and Excel files
"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any


class QAReporter:
    """
    Handles all reporting functionality for the QA Agent
    - Console logging with colored output
    - Excel report generation
    - Test result tracking
    """
    
    def __init__(self, report_dir: str = "reports"):
        """
        Initialize the reporter
        
        Args:
            report_dir: Directory to save reports
        """
        self.report_dir = report_dir
        self.test_results = []
        self.bug_count = 0
        self.pass_count = 0
        self.fail_count = 0
        
        # Create reports directory if it doesn't exist
        os.makedirs(report_dir, exist_ok=True)
    
    def log_test_result(self, test_case: str, result: str, details: str, 
                       test_type: str = "UI", severity: str = "Medium"):
        """
        Log a test result
        
        Args:
            test_case: Name of the test case
            result: PASS, FAIL, or BUG
            details: Detailed description of the result
            test_type: Type of test (UI, API, Form, Navigation, etc.)
            severity: Bug severity (Low, Medium, High, Critical)
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        test_result = {
            "Test Case": test_case,
            "Result": result,
            "Details": details,
            "Test Type": test_type,
            "Severity": severity,
            "Timestamp": timestamp
        }
        
        self.test_results.append(test_result)
        
        # Update counters
        if result == "PASS":
            self.pass_count += 1
        elif result == "FAIL":
            self.fail_count += 1
        elif result == "BUG":
            self.bug_count += 1
        
        # Console output with colors (basic implementation)
        status_color = self._get_status_color(result)
        print(f"{status_color}[{result}]{self._reset_color()} {test_case}")
        print(f"  Type: {test_type} | Severity: {severity}")
        print(f"  Details: {details}")
        print(f"  Time: {timestamp}")
        print("-" * 80)
    
    def _get_status_color(self, status: str) -> str:
        """Get color code for status (basic implementation)"""
        colors = {
            "PASS": "\033[92m",  # Green
            "FAIL": "\033[91m",  # Red
            "BUG": "\033[93m",   # Yellow
        }
        return colors.get(status, "\033[0m")  # Default to no color
    
    def _reset_color(self) -> str:
        """Reset color to default"""
        return "\033[0m"
    
    def generate_excel_report(self, filename: str = None, tested_host: str = None) -> str:
        """
        Generate Excel report with all test results and enhanced formatting
        
        Args:
            filename: Custom filename for the report
            
        Returns:
            Path to the generated report file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_host = (tested_host or "localhost").replace("/", "-") if tested_host else "localhost"
            filename = f"{safe_host}_qa_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.report_dir, filename)
        
        # Create DataFrame from test results
        df = pd.DataFrame(self.test_results)
        
        # Add enhanced formatting
        df = self._enhance_dataframe_formatting(df)
        
        # Create Excel writer with multiple sheets
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Main test results sheet with formatting
            df.to_excel(writer, sheet_name='Test Results', index=False)
            
            # Apply conditional formatting to the main sheet
            self._apply_excel_formatting(writer, 'Test Results', df)
            
            # Enhanced Summary sheet
            summary_data = self._create_enhanced_summary()
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Test Type Analysis
            test_type_analysis = self._create_test_type_analysis(df)
            test_type_df = pd.DataFrame(test_type_analysis)
            test_type_df.to_excel(writer, sheet_name='Test Type Analysis', index=False)
            
            # Severity Analysis
            severity_analysis = self._create_severity_analysis(df)
            severity_df = pd.DataFrame(severity_analysis)
            severity_df.to_excel(writer, sheet_name='Severity Analysis', index=False)
            
            # Bugs only sheet with priority sorting
            bugs_df = df[df['Result'] == 'BUG'].copy()
            if not bugs_df.empty:
                bugs_df = bugs_df.sort_values(['Severity', 'Test Type'], ascending=[False, True])
                bugs_df.to_excel(writer, sheet_name='Bugs Found', index=False)
            
            # Security Analysis
            security_df = df[df['Test Type'] == 'Security'].copy()
            if not security_df.empty:
                security_df.to_excel(writer, sheet_name='Security Tests', index=False)
            
            # Performance Analysis
            performance_df = df[df['Test Type'] == 'Performance'].copy()
            if not performance_df.empty:
                performance_df.to_excel(writer, sheet_name='Performance Tests', index=False)

            # Recommendations sheet based on findings
            recs = self._create_recommendations()
            pd.DataFrame(recs).to_excel(writer, sheet_name='Recommendations', index=False)

            # Add simple charts to Summary via openpyxl after save
            workbook = writer.book
            # Create chart for Pass/Fail/Bug in Summary
            try:
                from openpyxl.chart import PieChart, Reference, BarChart
                # Insert a small table in Summary for chart
                summary_ws = workbook['Summary']
                # Append counts at the bottom
                start_row = summary_ws.max_row + 3
                summary_ws.cell(row=start_row, column=1, value='Status')
                summary_ws.cell(row=start_row, column=2, value='Count')
                counts = {
                    'PASS': self.pass_count,
                    'FAIL': self.fail_count,
                    'BUG': self.bug_count
                }
                r = start_row + 1
                for k, v in counts.items():
                    summary_ws.cell(row=r, column=1, value=k)
                    summary_ws.cell(row=r, column=2, value=v)
                    r += 1

                pie = PieChart()
                pie.add_data(Reference(summary_ws, min_col=2, min_row=start_row+1, max_row=r-1), titles_from_data=False)
                pie.set_categories(Reference(summary_ws, min_col=1, min_row=start_row+1, max_row=r-1))
                pie.title = 'Results Distribution'
                summary_ws.add_chart(pie, f'E{start_row}')

                # Bar chart for Test Types
                if 'Test Type Analysis' in workbook.sheetnames:
                    tta = workbook['Test Type Analysis']
                    bar = BarChart()
                    # Assuming headers: Test Type | Total | Passed | Failed | Bugs | Pass Rate (%)
                    data_ref = Reference(tta, min_col=2, min_row=1, max_col=4, max_row=tta.max_row)
                    cats_ref = Reference(tta, min_col=1, min_row=2, max_row=tta.max_row)
                    bar.add_data(data_ref, titles_from_data=True)
                    bar.set_categories(cats_ref)
                    bar.title = 'Counts by Test Type'
                    tta.add_chart(bar, 'H2')
            except Exception:
                # Chart creation is best-effort; ignore if environment lacks chart modules
                pass
        
        print(f"\n📊 Enhanced Excel report generated: {filepath}")
        return filepath
    
    def _enhance_dataframe_formatting(self, df):
        """Enhance DataFrame with better formatting and visual indicators"""
        if df.empty:
            return df
        
        # Add professional indicators for results
        def format_result(result):
            if result == 'PASS':
                return 'PASS'
            elif result == 'FAIL':
                return 'FAIL'
            elif result == 'BUG':
                return 'BUG'
            return result
        
        # Add professional indicators for severity
        def format_severity(severity):
            if severity == 'Critical':
                return 'CRITICAL'
            elif severity == 'High':
                return 'HIGH'
            elif severity == 'Medium':
                return 'MEDIUM'
            elif severity == 'Low':
                return 'LOW'
            return severity
        
        # Add professional indicators for test types
        def format_test_type(test_type):
            return test_type
        
        # Apply formatting
        df['Result'] = df['Result'].apply(format_result)
        df['Severity'] = df['Severity'].apply(format_severity)
        df['Test Type'] = df['Test Type'].apply(format_test_type)
        
        # Add status summary
        df['Status Summary'] = df.apply(lambda row: 
            f"{row['Result']} | {row['Severity']} | {row['Test Type']}", axis=1)
        
        return df
    
    def _apply_excel_formatting(self, writer, sheet_name, df):
        """Apply conditional formatting to Excel sheets"""
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.formatting.rule import CellIsRule
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define colors
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Apply conditional formatting to Result column
        for row in range(2, len(df) + 2):  # Skip header
            result_cell = worksheet[f'B{row}']
            if 'PASS' in str(result_cell.value):
                result_cell.fill = green_fill
            elif 'FAIL' in str(result_cell.value):
                result_cell.fill = red_fill
            elif 'BUG' in str(result_cell.value):
                result_cell.fill = yellow_fill
        
        # Apply formatting to Severity column
        for row in range(2, len(df) + 2):
            severity_cell = worksheet[f'E{row}']
            if 'CRITICAL' in str(severity_cell.value):
                severity_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                severity_cell.font = Font(bold=True, color='FFFFFF')
            elif 'HIGH' in str(severity_cell.value):
                severity_cell.fill = PatternFill(start_color='FF8E53', end_color='FF8E53', fill_type='solid')
                severity_cell.font = Font(bold=True, color='FFFFFF')
    
    def _create_enhanced_summary(self):
        """Create enhanced summary with more detailed metrics"""
        total_tests = len(self.test_results)
        pass_rate = (self.pass_count / total_tests * 100) if total_tests > 0 else 0
        
        # Calculate security percentage
        security_tests = [r for r in self.test_results if r['Test Type'] == 'Security']
        security_passed = sum(1 for r in security_tests if r['Result'] == 'PASS')
        security_percentage = (security_passed / len(security_tests) * 100) if security_tests else 100
        
        # Calculate test type breakdown
        test_types = {}
        for result in self.test_results:
            test_type = result['Test Type']
            if test_type not in test_types:
                test_types[test_type] = {'total': 0, 'passed': 0, 'failed': 0, 'bugs': 0}
            test_types[test_type]['total'] += 1
            if result['Result'] == 'PASS':
                test_types[test_type]['passed'] += 1
            elif result['Result'] == 'FAIL':
                test_types[test_type]['failed'] += 1
            elif result['Result'] == 'BUG':
                test_types[test_type]['bugs'] += 1
        
        return {
            'Metric': [
                'Total Tests Executed',
                'Tests Passed',
                'Tests Failed',
                'Bugs Found',
                'Overall Pass Rate (%)',
                'Security Tests',
                'Security Pass Rate (%)',
                'Average Execution Time',
                'Test Coverage Score',
                'Critical Issues',
                'High Priority Issues',
                'Medium Priority Issues',
                'Low Priority Issues'
            ],
            'Value': [
                total_tests,
                self.pass_count,
                self.fail_count,
                self.bug_count,
                f"{pass_rate:.1f}%",
                len(security_tests),
                f"{security_percentage:.1f}%",
                "N/A",  # Will be calculated if execution time is available
                f"{pass_rate:.1f}%",  # Using pass rate as coverage score
                sum(1 for r in self.test_results if r['Severity'] == 'Critical'),
                sum(1 for r in self.test_results if r['Severity'] == 'High'),
                sum(1 for r in self.test_results if r['Severity'] == 'Medium'),
                sum(1 for r in self.test_results if r['Severity'] == 'Low')
            ]
        }
    
    def _create_test_type_analysis(self, df):
        """Create detailed test type analysis"""
        if df.empty:
            return {'Test Type': [], 'Total': [], 'Passed': [], 'Failed': [], 'Bugs': [], 'Pass Rate (%)': []}
        
        test_type_stats = {}
        for _, row in df.iterrows():
            test_type = row['Test Type']
            
            if test_type not in test_type_stats:
                test_type_stats[test_type] = {'total': 0, 'passed': 0, 'failed': 0, 'bugs': 0}
            
            test_type_stats[test_type]['total'] += 1
            if 'PASS' in str(row['Result']):
                test_type_stats[test_type]['passed'] += 1
            elif 'FAIL' in str(row['Result']):
                test_type_stats[test_type]['failed'] += 1
            elif 'BUG' in str(row['Result']):
                test_type_stats[test_type]['bugs'] += 1
        
        analysis = []
        for test_type, stats in test_type_stats.items():
            pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
            analysis.append({
                'Test Type': test_type,
                'Total': stats['total'],
                'Passed': stats['passed'],
                'Failed': stats['failed'],
                'Bugs': stats['bugs'],
                'Pass Rate (%)': f"{pass_rate:.1f}%"
            })
        
        return analysis
    
    def _create_severity_analysis(self, df):
        """Create detailed severity analysis"""
        if df.empty:
            return {'Severity': [], 'Count': [], 'Percentage (%)': [], 'Priority': []}
        
        severity_counts = {}
        for _, row in df.iterrows():
            severity = row['Severity']
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        total = sum(severity_counts.values())
        analysis = []
        
        priority_order = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}
        
        for severity, count in sorted(severity_counts.items(), key=lambda x: priority_order.get(x[0], 5)):
            percentage = (count / total * 100) if total > 0 else 0
            priority = "High" if severity in ['Critical', 'High'] else "Medium" if severity == 'Medium' else "Low"
            
            analysis.append({
                'Severity': severity,
                'Count': count,
                'Percentage (%)': f"{percentage:.1f}%",
                'Priority': priority
            })
        
        return analysis

    def _create_recommendations(self):
        """Generate human-friendly recommendations based on results."""
        recs = []
        total = len(self.test_results)
        if total == 0:
            recs.append({'Recommendation': 'No test results to analyze', 'Priority': 'Low'})
            return recs
        pass_rate = (self.pass_count / total) * 100 if total else 0
        if pass_rate < 80:
            recs.append({'Recommendation': 'Increase unit and integration test coverage to improve pass rate', 'Priority': 'High'})
        if self.bug_count > 0:
            recs.append({'Recommendation': 'Triage and fix open BUG results; prioritize Critical/High severities', 'Priority': 'High'})
        security_bugs = [r for r in self.test_results if r['Test Type'] == 'Security' and r['Result'] in ('FAIL', 'BUG')]
        if security_bugs:
            recs.append({'Recommendation': 'Address security test failures urgently and add regression tests', 'Priority': 'Critical'})
        flaky = [r for r in self.test_results if 'timeout' in str(r['Details']).lower()]
        if flaky:
            recs.append({'Recommendation': 'Investigate flaky tests (timeouts). Add waits/retries and stabilize selectors', 'Priority': 'Medium'})
        return recs
    
    def print_summary(self):
        """Print test execution summary to console"""
        print("\n" + "="*80)
        print("🔍 QA AGENT TEST SUMMARY")
        print("="*80)
        print(f"Total Tests Executed: {len(self.test_results)}")
        print(f"✅ Passed: {self.pass_count}")
        print(f"❌ Failed: {self.fail_count}")
        print(f"🐛 Bugs Found: {self.bug_count}")
        
        if self.test_results:
            pass_rate = (self.pass_count / len(self.test_results)) * 100
            print(f"📈 Pass Rate: {pass_rate:.1f}%")
        
        print("="*80)
        
        if self.bug_count > 0:
            print(f"\n⚠️  {self.bug_count} bugs found! Check the Excel report for details.")
        else:
            print("\n🎉 No bugs found! All tests passed.")
    
    def get_bug_summary(self) -> List[Dict[str, Any]]:
        """
        Get summary of all bugs found
        
        Returns:
            List of bug dictionaries
        """
        return [result for result in self.test_results if result['Result'] == 'BUG']
    
    def clear_results(self):
        """Clear all test results (useful for running multiple test suites)"""
        self.test_results = []
        self.bug_count = 0
        self.pass_count = 0
        self.fail_count = 0


# Global reporter instance
reporter = QAReporter()
